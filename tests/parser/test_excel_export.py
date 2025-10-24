from uuid import uuid4

import openpyxl
import pandas as pd
import pytest

from parser.excel_export import export_offers_xlsx, pivot_offers_for_export
from parser.types import OfferRow


@pytest.fixture
def sample_offers() -> list[OfferRow]:
    """Provides a list of sample OfferRow objects for testing."""
    return [
        # Article 1 has 3 offers
        OfferRow(
            b="BrandA",
            a="ART1",
            name="Part A1",
            price=10.0,
            quantity=100,
            provider="WH1",
            delivery=1,
            is_analog=False,
            rating=90.0,
        ),
        OfferRow(
            b="BrandA",
            a="ART1",
            name="Part A2",
            price=12.0,
            quantity=50,
            provider="WH2",
            delivery=2,
            is_analog=False,
            rating=80.0,
        ),
        OfferRow(
            b="BrandA",
            a="ART1",
            name="Part A3",
            price=11.0,
            quantity=200,
            provider="WH3",
            delivery=3,
            is_analog=False,
            rating=85.0,
        ),
        # Article 2 has 1 offer
        OfferRow(
            b="BrandB",
            a="ART2",
            name="Part B",
            price=50.0,
            quantity=10,
            provider="WH4",
            delivery=5,
            is_analog=False,
            rating=95.0,
        ),
    ]


def test_pivot_offers_for_export(sample_offers):
    """Verify that the pivot function correctly transforms long data to wide format."""
    df_wide = pivot_offers_for_export(sample_offers)

    assert len(df_wide) == 2  # One row for each unique article
    assert "price 3" in df_wide.columns
    assert "price 11" not in df_wide.columns

    # Check data for ART1
    art1_row = df_wide[df_wide["article"] == "ART1"].iloc[0]
    # The offers should be sorted by price, so 10.0, 11.0, 12.0
    assert art1_row["price 1"] == 10.0
    assert art1_row["supplier 1"] == "WH1"
    assert art1_row["price 2"] == 11.0
    assert art1_row["supplier 2"] == "WH3"
    assert art1_row["price 3"] == 12.0
    assert art1_row["supplier 3"] == "WH2"

    # Check data for ART2
    art2_row = df_wide[df_wide["article"] == "ART2"].iloc[0]
    assert art2_row["price 1"] == 50.0
    assert art2_row["supplier 1"] == "WH4"
    assert pd.isna(art2_row["price 2"])


def test_export_offers_xlsx(sample_offers, tmp_path):
    """Verify that the Excel export function creates a correctly formatted file."""
    run_id = uuid4()
    source = "stparts"
    df_wide = pivot_offers_for_export(sample_offers)

    file_path = export_offers_xlsx(run_id, source, df_wide, tmp_path)

    assert file_path.exists()
    assert file_path.name == f"stparts_{run_id}.xlsx"

    # --- Verify Excel content and formatting --- #
    wb = openpyxl.load_workbook(file_path)
    ws = wb[source]

    # 1. Check merged header
    merged_header = ws["A1"].value
    assert merged_header == f"Source: {source}"
    # Check the merge range directly
    assert any(cr.coord == "A1:AZ1" for cr in ws.merged_cells.ranges)

    # 2. Check data headers (should be in row 2)
    assert ws["A2"].value == "brand"
    assert ws["C2"].value == "price 1"

    # 3. Check data content
    assert ws["B3"].value == "ART1"
    assert ws["C3"].value == 10.0

    # 4. Check frozen panes
    assert ws.freeze_panes == "A3"

    # 5. Check number format on a price column
    price_cell = ws["C3"]
    assert price_cell.number_format == "#,##0.00"


def test_export_offers_xlsx_empty_dataframe(tmp_path):
    """Verify that an empty DataFrame results in a valid Excel file with only headers."""
    run_id = uuid4()
    source = "stparts"
    df_empty = pivot_offers_for_export([])  # This creates an empty df with columns

    file_path = export_offers_xlsx(run_id, source, df_empty, tmp_path)

    assert file_path.exists()
    wb = openpyxl.load_workbook(file_path)
    ws = wb[source]

    # Check that the merged header is there
    assert ws["A1"].value == f"Source: {source}"
    # Check that the data headers are there
    assert ws["A2"].value == "brand"
    assert ws["B2"].value == "article"
    # Check that there is no data in the first data row
    assert ws["A3"].value is None
