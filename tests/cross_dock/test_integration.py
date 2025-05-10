"""
Integration tests for cross-dock functionality.

This module contains integration tests for the cross-dock functionality,
testing the full workflow from file input to Excel output.
"""

import os
import tempfile
from unittest import mock

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

# Import will fail until we implement the code
from cross_dock.services.excel_service import process_cross_dock_data_from_file


class TestIntegration:
    """Integration test suite for cross-dock functionality."""

    @mock.patch('cross_dock.services.clickhouse_service.query_supplier_data')
    def test_process_cross_dock_data_from_file(self, mock_query_supplier_data):
        """Test processing cross-dock data from a file."""
        # Mock the query_supplier_data function
        mock_query_supplier_data.return_value = pd.DataFrame({
            "price": [100.50, 120.75, 90.25],
            "quantity": [5, 10, 3],
            "supplier_name": ["Supplier A", "Supplier B", "Supplier C"]
        })

        # Create a temporary input file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file_path = temp_file.name

        # Create a test input Excel file
        wb = Workbook()
        sheet = wb.active

        # Add headers
        sheet.cell(row=1, column=1, value="Бренд")
        sheet.cell(row=1, column=2, value="Артикул")

        # Add data
        sheet.cell(row=2, column=1, value="HYUNDAI/KIA/MOBIS")
        sheet.cell(row=2, column=2, value="223112e100")
        sheet.cell(row=3, column=1, value="VAG")
        sheet.cell(row=3, column=2, value="000915105cd")

        # Save the workbook
        wb.save(temp_file_path)

        try:
            # Process the file
            output_file_path = process_cross_dock_data_from_file(temp_file_path, "emex")

            # Check that the output file exists
            assert os.path.exists(output_file_path)

            # Load the output workbook and check its contents
            wb = load_workbook(output_file_path)
            sheet = wb.active

            # Check headers
            headers = ["SKU", "Бренд", "Артикул", "Лучшая цена 1", "Количество 1", "Название поставщика 1",
                    "Лучшая цена 2", "Количество 2", "Название поставщика 2",
                    "Лучшая цена 3", "Количество 3", "Название поставщика 3"]
            for col_num, header in enumerate(headers, start=1):
                assert sheet.cell(row=1, column=col_num).value == header

            # Check first row data
            assert sheet.cell(row=2, column=1).value == "HYUNDAI/KIA/MOBIS|223112e100"  # SKU
            assert sheet.cell(row=2, column=2).value == "HYUNDAI/KIA/MOBIS"  # Brand
            assert sheet.cell(row=2, column=3).value == "223112e100"  # Article
            assert sheet.cell(row=2, column=4).value == 100.50  # Price 1
            assert sheet.cell(row=2, column=5).value == 5  # Quantity 1
            assert sheet.cell(row=2, column=6).value == "Supplier A"  # Supplier 1

            # Clean up
            os.unlink(output_file_path)
        finally:
            # Clean up the input file
            os.unlink(temp_file_path)
