"""
Tests for Excel service.

This module contains tests for the Excel service functions that are used
for processing cross-dock data and generating Excel files.
"""

import os
import tempfile
from unittest import mock

import pandas as pd
from openpyxl import load_workbook

from cross_dock.services.excel_service import process_cross_dock_data


class TestExcelService:
    """Test suite for Excel service functions."""

    @mock.patch("cross_dock.services.clickhouse_service.query_supplier_data")
    def test_process_cross_dock_data(self, mock_query_supplier_data):
        """Test processing cross-dock data and generating an Excel file."""
        mock_query_supplier_data.return_value = pd.DataFrame(
            {
                "price": [100.50, 120.75, 90.25],
                "quantity": [5, 10, 3],
                "supplier_name": ["Supplier A", "Supplier B", "Supplier C"],
            }
        )

        test_data = [
            {"Бренд": "HYUNDAI/KIA/MOBIS", "Артикул": "223112e100"},
            {"Бренд": "VAG", "Артикул": "000915105cd"},
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch("django.conf.settings.MEDIA_ROOT", temp_dir):
                with mock.patch("django.conf.settings.MEDIA_URL", "/media/"):
                    progress, file_url = process_cross_dock_data(test_data, "Группа для проценки ТРЕШКА")
                    assert progress == "100%"

                    file_path = os.path.join(temp_dir, "exports", os.path.basename(file_url))
                    assert os.path.exists(file_path)

                    wb = load_workbook(file_path)
                    sheet = wb.active

                    headers = [
                        "SKU",
                        "Бренд",
                        "Артикул",
                        "Лучшая цена 1",
                        "Количество 1",
                        "Название поставщика 1",
                        "Лучшая цена 2",
                        "Количество 2",
                        "Название поставщика 2",
                        "Лучшая цена 3",
                        "Количество 3",
                        "Название поставщика 3",
                    ]
                    for col_num, header in enumerate(headers, start=1):
                        assert sheet.cell(row=1, column=col_num).value == header

                    assert sheet.cell(row=2, column=1).value == "HYUNDAI/KIA/MOBIS|223112e100"
                    assert sheet.cell(row=2, column=2).value == "HYUNDAI/KIA/MOBIS"
                    assert sheet.cell(row=2, column=3).value == "223112e100"
                    assert sheet.cell(row=2, column=4).value == 100.50
                    assert sheet.cell(row=2, column=5).value == 5
                    assert sheet.cell(row=2, column=6).value == "Supplier A"

    @mock.patch("cross_dock.services.clickhouse_service.query_supplier_data")
    def test_process_cross_dock_data_empty_results(self, mock_query_supplier_data):
        """Test handling of empty query results."""
        mock_query_supplier_data.return_value = pd.DataFrame(columns=["price", "quantity", "supplier_name"])

        test_data = [{"Бренд": "BRAND", "Артикул": "ARTICLE"}]

        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch("django.conf.settings.MEDIA_ROOT", temp_dir):
                with mock.patch("django.conf.settings.MEDIA_URL", "/media/"):
                    progress, file_url = process_cross_dock_data(test_data, "Группа для проценки ТРЕШКА")
                    assert progress == "100%"

                    file_path = os.path.join(temp_dir, "exports", os.path.basename(file_url))
                    assert os.path.exists(file_path)

                    wb = load_workbook(file_path)
                    sheet = wb.active

                    # Verify empty cells are handled properly
                    assert sheet.cell(row=2, column=4).value is None
                    assert sheet.cell(row=2, column=5).value is None
                    assert sheet.cell(row=2, column=6).value is None

    @mock.patch("cross_dock.services.clickhouse_service.query_supplier_data")
    def test_process_cross_dock_data_exception(self, mock_query_supplier_data):
        """Test error handling during data processing."""
        mock_query_supplier_data.side_effect = Exception("Test exception")

        test_data = [{"Бренд": "BRAND", "Артикул": "ARTICLE"}]

        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch("django.conf.settings.MEDIA_ROOT", temp_dir):
                with mock.patch("django.conf.settings.MEDIA_URL", "/media/"):
                    progress, file_url = process_cross_dock_data(test_data, "Группа для проценки ТРЕШКА")
                    assert progress == "100%"

                    file_path = os.path.join(temp_dir, "exports", os.path.basename(file_url))
                    assert os.path.exists(file_path)

                    wb = load_workbook(file_path)
                    sheet = wb.active

                    # Verify error handling by checking for empty cells
                    assert sheet.cell(row=2, column=4).value is None
                    assert sheet.cell(row=2, column=5).value is None
                    assert sheet.cell(row=2, column=6).value is None
